import argparse
import hashlib
import json
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from app.db.session import SessionLocal
from app.models.import_batch import ImportBatch
from app.models.import_row import ImportRow

DEFAULT_SHEET = "Form responses"
SOURCE_COLUMN_COUNT = 18


@dataclass(frozen=True)
class ParsedRow:
    row_number: int
    source_record_id: str | None
    row_fingerprint: str
    raw_data: dict[str, object]
    validation_issues: list[str]


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def json_safe_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if value is None or isinstance(
        value,
        (str, int, float, bool),
    ):
        return value

    return str(value)


def calculate_file_sha256(file_path: Path) -> str:
    digest = hashlib.sha256()

    with file_path.open("rb") as workbook_file:
        while chunk := workbook_file.read(1024 * 1024):
            digest.update(chunk)

    return digest.hexdigest()


def calculate_row_fingerprint(
    raw_data: dict[str, object],
) -> str:
    encoded_row = json.dumps(
        raw_data,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")

    return hashlib.sha256(encoded_row).hexdigest()


def validate_raw_row(
    raw_data: dict[str, object],
) -> list[str]:
    issues = []

    if is_blank(raw_data.get("B")):
        issues.append("missing delivery date")

    if is_blank(raw_data.get("C")):
        issues.append("missing food recipient")

    if is_blank(raw_data.get("F")):
        issues.append("missing vehicle")

    if is_blank(raw_data.get("P")):
        issues.append("missing recorded total pounds")

    if is_blank(raw_data.get("R")):
        issues.append("missing source record id")

    return issues


def parse_workbook(
    file_path: Path,
    sheet_name: str = DEFAULT_SHEET,
) -> list[ParsedRow]:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"workbook does not contain sheet {sheet_name!r}")

        worksheet = workbook[sheet_name]
        parsed_rows = []

        rows = worksheet.iter_rows(
            min_row=2,
            max_col=SOURCE_COLUMN_COUNT,
            values_only=True,
        )

        for row_number, values in enumerate(rows, start=2):
            raw_data = {
                get_column_letter(column_number): json_safe_value(value)
                for column_number, value in enumerate(
                    values,
                    start=1,
                )
            }

            raw_source_id = raw_data.get("R")
            source_record_id = None if is_blank(raw_source_id) else str(raw_source_id).strip()
            validation_issues = validate_raw_row(raw_data)

            parsed_rows.append(
                ParsedRow(
                    row_number=row_number,
                    source_record_id=source_record_id,
                    row_fingerprint=calculate_row_fingerprint(raw_data),
                    raw_data=raw_data,
                    validation_issues=validation_issues,
                )
            )

        return parsed_rows
    finally:
        workbook.close()


def build_summary(
    file_path: Path,
    parsed_rows: list[ParsedRow],
) -> dict[str, object]:
    issue_counts = Counter(
        issue for parsed_row in parsed_rows for issue in parsed_row.validation_issues
    )
    review_rows = sum(bool(parsed_row.validation_issues) for parsed_row in parsed_rows)

    return {
        "file": file_path.name,
        "file_sha256": calculate_file_sha256(file_path),
        "sheet": DEFAULT_SHEET,
        "total_rows": len(parsed_rows),
        "clean_rows": len(parsed_rows) - review_rows,
        "review_rows": review_rows,
        "issue_counts": dict(sorted(issue_counts.items())),
    }


def stage_rows(
    file_path: Path,
    parsed_rows: list[ParsedRow],
    summary: dict[str, object],
) -> str:
    file_sha256 = str(summary["file_sha256"])

    with SessionLocal.begin() as database:
        existing_batch = database.scalar(
            select(ImportBatch).where(ImportBatch.file_sha256 == file_sha256)
        )

        if existing_batch is not None:
            raise ValueError(f"this workbook has already been staged as batch {existing_batch.id}")

        batch = ImportBatch(
            file_name=file_path.name,
            file_sha256=file_sha256,
            source_sheet=DEFAULT_SHEET,
            status="staged",
            total_rows=len(parsed_rows),
            review_rows=int(summary["review_rows"]),
        )
        database.add(batch)
        database.flush()

        database.add_all(
            [
                ImportRow(
                    batch_id=batch.id,
                    sheet_name=DEFAULT_SHEET,
                    row_number=parsed_row.row_number,
                    source_record_id=(parsed_row.source_record_id),
                    row_fingerprint=(parsed_row.row_fingerprint),
                    raw_data=parsed_row.raw_data,
                    status=("needs_review" if parsed_row.validation_issues else "pending"),
                    validation_issues=(parsed_row.validation_issues),
                )
                for parsed_row in parsed_rows
            ]
        )

    return str(batch.id)


def main() -> None:
    parser = argparse.ArgumentParser(description="Stage historical Veggie Rescue deliveries.")
    parser.add_argument(
        "--file",
        required=True,
        type=Path,
        help="Path to the exported Google Sheets workbook.",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Write staged rows to PostgreSQL.",
    )

    arguments = parser.parse_args()
    file_path = arguments.file.resolve()

    if not file_path.is_file():
        parser.error(f"file does not exist: {file_path}")

    parsed_rows = parse_workbook(file_path)
    summary = build_summary(file_path, parsed_rows)

    if arguments.commit:
        summary["batch_id"] = stage_rows(
            file_path,
            parsed_rows,
            summary,
        )
        summary["committed"] = True
    else:
        summary["committed"] = False

    print(
        json.dumps(
            summary,
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
